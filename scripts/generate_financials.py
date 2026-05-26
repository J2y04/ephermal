"""
Ephermal Financial Model — generates Ephermal_Financials.xlsx
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import os

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY      = "0D0F1A"
PURPLE    = "6366F1"
TEAL      = "06D6C7"
GREEN     = "34D399"
YELLOW    = "FBBF24"
RED       = "F87171"
WHITE     = "FFFFFF"
LIGHT_BG  = "F0F0FF"
MID_BG    = "1A1D2E"
BORDER_C  = "2D3150"
MUTED     = "8892B0"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

def border(style="thin", color=BORDER_C):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color=BORDER_C):
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def money(val):
    return f"${val:,.0f}"

def pct(val):
    return f"{val:.0f}%"

# ── Workbook ────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Overview / Summary
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("📊 Overview")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 2
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 18
ws1.column_dimensions["F"].width = 18
ws1.column_dimensions["G"].width = 2

# Background
for row in ws1.iter_rows(min_row=1, max_row=50, min_col=1, max_col=7):
    for cell in row:
        cell.fill = fill(NAVY)

# Title block
ws1.merge_cells("B1:F1")
ws1["B1"] = "EPHERMAL — FINANCIAL MODEL"
ws1["B1"].font = Font(bold=True, color=TEAL, size=16, name="Calibri")
ws1["B1"].alignment = center()
ws1["B1"].fill = fill(NAVY)

ws1.merge_cells("B2:F2")
ws1["B2"] = "Pay-As-You-Go · Higgsfield Marketing Studio · May 2026"
ws1["B2"].font = Font(color=MUTED, size=10, name="Calibri", italic=True)
ws1["B2"].alignment = center()

ws1.row_dimensions[1].height = 36
ws1.row_dimensions[2].height = 20
ws1.row_dimensions[3].height = 12

# ── KPI cards ──────────────────────────────────────────────────────────────────
def kpi_card(ws, col_letter, row, label, value, color, note=""):
    c1 = ws[f"{col_letter}{row}"]
    c1.value = label
    c1.font = Font(color=MUTED, size=9, name="Calibri", bold=True)
    c1.alignment = center()
    c1.fill = fill("11142B")

    c2 = ws[f"{col_letter}{row+1}"]
    c2.value = value
    c2.font = Font(color=color, size=20, name="Calibri", bold=True)
    c2.alignment = center()
    c2.fill = fill("11142B")

    if note:
        c3 = ws[f"{col_letter}{row+2}"]
        c3.value = note
        c3.font = Font(color=MUTED, size=8, name="Calibri", italic=True)
        c3.alignment = center()
        c3.fill = fill("11142B")

ws1.row_dimensions[4].height = 18
ws1.row_dimensions[5].height = 32
ws1.row_dimensions[6].height = 18
ws1.row_dimensions[7].height = 12

kpi_card(ws1, "B", 4, "BREAK-EVEN", "1 customer", GREEN, "covers infra")
kpi_card(ws1, "C", 4, "AVG MARGIN", "93%", TEAL, "at 50+ customers")
kpi_card(ws1, "D", 4, "INFRA COST", "$30–80/mo", YELLOW, "fixed baseline")
kpi_card(ws1, "E", 4, "PAYG RATE", "$0.28/video", PURPLE, "18 cr / $1")
kpi_card(ws1, "F", 4, "ULTRA RATE", "$0.165/video", GREEN, "switch at 4+ Growth")

ws1.row_dimensions[8].height = 12

# ── Plan comparison ────────────────────────────────────────────────────────────
def hdr(ws, row, cols, labels, bg=MID_BG):
    ws.row_dimensions[row].height = 22
    for col, label in zip(cols, labels):
        c = ws[f"{col}{row}"]
        c.value = label
        c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        c.alignment = center()
        c.fill = fill(bg)
        c.border = border()

# Plan overview header
ws1.row_dimensions[9].height = 22
ws1.merge_cells("B9:F9")
ws1["B9"] = "SUBSCRIPTION PLAN SUMMARY"
ws1["B9"].font = Font(bold=True, color=PURPLE, size=11, name="Calibri")
ws1["B9"].alignment = left()
ws1["B9"].fill = fill("11142B")

hdr(ws1, 10, ["B","C","D","E","F"],
    ["Plan", "Price/mo", "UGC/mo", "Stores", "Target Customer"])

plans = [
    ("Starter", "$89", "15 videos", "1 store",       "Solo DTC founders"),
    ("Growth",  "$159","75 videos", "Up to 3 stores", "Growing brands"),
    ("Scale",   "$349","350 videos","Unlimited",      "Agencies / multi-brand"),
]
plan_colors = [MUTED, PURPLE, TEAL]
for i, (name, price, ugc, stores, target) in enumerate(plans):
    r = 11 + i
    ws1.row_dimensions[r].height = 20
    row_bg = "13162A" if i % 2 == 0 else "0F1120"
    for col, val in zip(["B","C","D","E","F"], [name, price, ugc, stores, target]):
        c = ws1[f"{col}{r}"]
        c.value = val
        c.font = Font(color=plan_colors[i] if col == "B" else WHITE, size=10,
                      name="Calibri", bold=(col == "B"))
        c.alignment = center()
        c.fill = fill(row_bg)
        c.border = border()

ws1.row_dimensions[14].height = 12

# ── Higgsfield plan comparison ─────────────────────────────────────────────────
ws1.merge_cells("B15:F15")
ws1["B15"] = "HIGGSFIELD PLAN COMPARISON"
ws1["B15"].font = Font(bold=True, color=TEAL, size=11, name="Calibri")
ws1["B15"].alignment = left()
ws1["B15"].fill = fill("11142B")

hdr(ws1, 16, ["B","C","D","E","F"],
    ["Option", "Cost/mo", "Credits/mo", "Cost/video", "Best for"])

hig_plans = [
    ("Pay-As-You-Go", "Varies",  "18 cr / $1",  "$0.28",  "0–5 customers"),
    ("PLUS Annual",   "$39",     "1,000 cr/mo",  "$0.20",  "5–15 customers"),
    ("ULTRA Annual",  "$99",     "3,000 cr/mo",  "$0.165", "15+ customers"),
]
hig_colors = [YELLOW, PURPLE, GREEN]
for i, (opt, cost, cr, cpv, best) in enumerate(hig_plans):
    r = 17 + i
    ws1.row_dimensions[r].height = 20
    row_bg = "13162A" if i % 2 == 0 else "0F1120"
    for col, val in zip(["B","C","D","E","F"], [opt, cost, cr, cpv, best]):
        c = ws1[f"{col}{r}"]
        c.value = val
        c.font = Font(color=hig_colors[i] if col == "B" else WHITE, size=10,
                      name="Calibri", bold=(col == "B"))
        c.alignment = center()
        c.fill = fill(row_bg)
        c.border = border()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — P&L Projections (PAYG)
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("💰 P&L Projections")
ws2.sheet_view.showGridLines = False

col_widths = [2, 18, 14, 14, 14, 16, 16, 14, 16, 2]
col_letters = [get_column_letter(i+1) for i in range(len(col_widths))]
for ltr, w in zip(col_letters, col_widths):
    ws2.column_dimensions[ltr].width = w

for row in ws2.iter_rows(min_row=1, max_row=60, min_col=1, max_col=10):
    for cell in row:
        cell.fill = fill(NAVY)

ws2.merge_cells("B1:I1")
ws2["B1"] = "P&L PROJECTIONS — PAY-AS-YOU-GO (Higgsfield)"
ws2["B1"].font = Font(bold=True, color=TEAL, size=14, name="Calibri")
ws2["B1"].alignment = center()
ws2.row_dimensions[1].height = 36

ws2.merge_cells("B2:I2")
ws2["B2"] = "Customer mix: 60% Starter ($89) / 30% Growth ($159) / 10% Scale ($349)   ·   $0.28/video PAYG"
ws2["B2"].font = Font(color=MUTED, size=9, name="Calibri", italic=True)
ws2["B2"].alignment = center()
ws2.row_dimensions[2].height = 18
ws2.row_dimensions[3].height = 10

hdr(ws2, 4, ["B","C","D","E","F","G","H","I"],
    ["Customers","Infra Cost","Higgsfield","Total Cost","Revenue","Gross Profit","Margin %","Break-even"],
    bg=MID_BG)

scenarios = [5, 10, 25, 50, 100, 250, 500]

def calc_row(n):
    starters = round(n * 0.60)
    growth   = round(n * 0.30)
    scale    = n - starters - growth

    infra = 30 if n <= 25 else 55 if n <= 100 else 80

    # PAYG @ $0.28/video
    hf = (starters * 15 + growth * 75 + scale * 350) * 0.28

    total_cost = infra + hf

    rev = starters * 89 + growth * 159 + scale * 349
    profit = rev - total_cost
    margin = profit / rev * 100 if rev > 0 else 0
    return infra, hf, total_cost, rev, profit, margin

for i, n in enumerate(scenarios):
    r = 5 + i
    ws2.row_dimensions[r].height = 22
    infra, hf, cost, rev, profit, margin = calc_row(n)
    row_bg = "13162A" if i % 2 == 0 else "0F1120"

    vals = [
        (str(n) + " users",   WHITE,  center()),
        (money(infra),         YELLOW, center()),
        (money(hf),            PURPLE, center()),
        (money(cost),          RED,    center()),
        (money(rev),           TEAL,   center()),
        (money(profit),        GREEN,  center()),
        (pct(margin),          GREEN,  center()),
        ("Month 1" if n <= 10 else f"~{max(1,round(cost/89))} customers", MUTED, center()),
    ]
    for col, (val, col_color, align) in zip(["B","C","D","E","F","G","H","I"], vals):
        c = ws2[f"{col}{r}"]
        c.value = val
        c.font = Font(color=col_color, size=10, name="Calibri",
                      bold=(col == "B" or col == "G"))
        c.alignment = align
        c.fill = fill(row_bg)
        c.border = border()

# Upgrade recommendation note
ws2.row_dimensions[13].height = 10
ws2.merge_cells("B14:I14")
ws2["B14"] = "⚡  Switch to ULTRA Annual ($99/mo) once you have 4+ Growth or 1+ Scale customers — saves ~$50–280/mo"
ws2["B14"].font = Font(color=YELLOW, size=10, name="Calibri", bold=True)
ws2["B14"].alignment = left()
ws2["B14"].fill = fill("1A1500")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Per-Customer Unit Economics
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🧮 Unit Economics")
ws3.sheet_view.showGridLines = False

col_w3 = [2, 26, 16, 16, 16, 16, 16, 2]
for i, w in enumerate(col_w3):
    ws3.column_dimensions[get_column_letter(i+1)].width = w

for row in ws3.iter_rows(min_row=1, max_row=60, min_col=1, max_col=8):
    for cell in row:
        cell.fill = fill(NAVY)

ws3.merge_cells("B1:G1")
ws3["B1"] = "UNIT ECONOMICS — PER CUSTOMER"
ws3["B1"].font = Font(bold=True, color=TEAL, size=14, name="Calibri")
ws3["B1"].alignment = center()
ws3.row_dimensions[1].height = 36
ws3.row_dimensions[2].height = 10

# ── PAYG section ───────────────────────────────────────────────────────────────
ws3.merge_cells("B3:G3")
ws3["B3"] = "PAY-AS-YOU-GO  ($0.28/video — 18 credits per $1)"
ws3["B3"].font = Font(bold=True, color=YELLOW, size=11, name="Calibri")
ws3["B3"].alignment = left()
ws3["B3"].fill = fill("1A1500")

hdr(ws3, 4, ["B","C","D","E","F","G"],
    ["Tier", "Revenue/mo", "UGC Videos", "Higgsfield Cost", "Gross Profit", "UGC Margin"])

payg_tiers = [
    ("Starter", 89,  15,  0.28),
    ("Growth",  159, 75,  0.28),
    ("Scale",   349, 350, 0.28),
]
tier_colors = [MUTED, PURPLE, TEAL]
for i, (name, rev, vids, rate) in enumerate(payg_tiers):
    r = 5 + i
    ws3.row_dimensions[r].height = 22
    hf_cost = vids * rate
    profit = rev - hf_cost
    ugc_margin = profit / rev * 100
    row_bg = "13162A" if i % 2 == 0 else "0F1120"
    for col, val, col_c in zip(
        ["B","C","D","E","F","G"],
        [name, money(rev), str(vids), money(hf_cost), money(profit), pct(ugc_margin)],
        [tier_colors[i], TEAL, WHITE, RED, GREEN, GREEN]
    ):
        c = ws3[f"{col}{r}"]
        c.value = val
        c.font = Font(color=col_c, size=10, name="Calibri", bold=(col in ["B","F","G"]))
        c.alignment = center()
        c.fill = fill(row_bg)
        c.border = border()

ws3.row_dimensions[8].height = 12

# ── ULTRA Annual section ────────────────────────────────────────────────────────
ws3.merge_cells("B9:G9")
ws3["B9"] = "ULTRA ANNUAL  ($0.165/video — 30.3 credits per $1)"
ws3["B9"].font = Font(bold=True, color=GREEN, size=11, name="Calibri")
ws3["B9"].alignment = left()
ws3["B9"].fill = fill("001A10")

hdr(ws3, 10, ["B","C","D","E","F","G"],
    ["Tier", "Revenue/mo", "UGC Videos", "Higgsfield Cost", "Gross Profit", "UGC Margin"])

ultra_tiers = [
    ("Starter", 89,  15,  0.165),
    ("Growth",  159, 75,  0.165),
    ("Scale",   349, 350, 0.165),
]
for i, (name, rev, vids, rate) in enumerate(ultra_tiers):
    r = 11 + i
    ws3.row_dimensions[r].height = 22
    hf_cost = vids * rate
    profit = rev - hf_cost
    ugc_margin = profit / rev * 100
    row_bg = "13162A" if i % 2 == 0 else "0F1120"
    for col, val, col_c in zip(
        ["B","C","D","E","F","G"],
        [name, money(rev), str(vids), money(hf_cost), money(profit), pct(ugc_margin)],
        [tier_colors[i], TEAL, WHITE, RED, GREEN, GREEN]
    ):
        c = ws3[f"{col}{r}"]
        c.value = val
        c.font = Font(color=col_c, size=10, name="Calibri", bold=(col in ["B","F","G"]))
        c.alignment = center()
        c.fill = fill(row_bg)
        c.border = border()

ws3.row_dimensions[14].height = 12

# ── Annual LTV ──────────────────────────────────────────────────────────────────
ws3.merge_cells("B15:G15")
ws3["B15"] = "ANNUAL LTV  (assuming 12-month retention)"
ws3["B15"].font = Font(bold=True, color=PURPLE, size=11, name="Calibri")
ws3["B15"].alignment = left()
ws3["B15"].fill = fill("100A1A")

hdr(ws3, 16, ["B","C","D","E","F","G"],
    ["Tier", "Monthly Rev", "Annual LTV", "Annual COGS (PAYG)", "Annual Profit", "Annual Margin"])

ltv_tiers = [
    ("Starter", 89,  15,  0.28),
    ("Growth",  159, 75,  0.28),
    ("Scale",   349, 350, 0.28),
]
for i, (name, rev, vids, rate) in enumerate(ltv_tiers):
    r = 17 + i
    ws3.row_dimensions[r].height = 22
    annual_rev  = rev * 12
    annual_cogs = vids * rate * 12
    annual_profit = annual_rev - annual_cogs
    annual_margin = annual_profit / annual_rev * 100
    row_bg = "13162A" if i % 2 == 0 else "0F1120"
    for col, val, col_c in zip(
        ["B","C","D","E","F","G"],
        [name, money(rev), money(annual_rev), money(annual_cogs), money(annual_profit), pct(annual_margin)],
        [tier_colors[i], TEAL, GREEN, RED, GREEN, GREEN]
    ):
        c = ws3[f"{col}{r}"]
        c.value = val
        c.font = Font(color=col_c, size=10, name="Calibri", bold=(col in ["B","D","E","F","G"]))
        c.alignment = center()
        c.fill = fill(row_bg)
        c.border = border()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Infrastructure Costs
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("🏗️ Infrastructure")
ws4.sheet_view.showGridLines = False

col_w4 = [2, 22, 20, 18, 22, 22, 2]
for i, w in enumerate(col_w4):
    ws4.column_dimensions[get_column_letter(i+1)].width = w

for row in ws4.iter_rows(min_row=1, max_row=50, min_col=1, max_col=7):
    for cell in row:
        cell.fill = fill(NAVY)

ws4.merge_cells("B1:F1")
ws4["B1"] = "INFRASTRUCTURE COST BREAKDOWN"
ws4["B1"].font = Font(bold=True, color=TEAL, size=14, name="Calibri")
ws4["B1"].alignment = center()
ws4.row_dimensions[1].height = 36
ws4.row_dimensions[2].height = 10

hdr(ws4, 3, ["B","C","D","E","F"],
    ["Service", "Role", "Free Tier", "Paid Tier", "Upgrade Trigger"])

infra = [
    ("Vercel",      "Landing page + Dashboard hosting",    "Hobby — Free",           "Pro — $20/mo",             "Custom domain / team"),
    ("n8n (Railway)","21 automation workflows",             "Self-hosted ~$7/mo",     "Add dyno ~$15/mo",         "High execution volume"),
    ("Clerk",       "Authentication",                       "Free up to 10k MAU",     "$0.02 per MAU after",      "~10,000 registered users"),
    ("Supabase",    "Database (integrations, creatives)",   "Free — 500MB / 50k rows","Pro — $25/mo",             "~100 active customers"),
    ("Higgsfield",  "UGC / Marketing Studio videos",        "8 credits (depleted)",   "PLUS $39 / ULTRA $99/mo",  "Day 1 — upgrade now"),
    ("OpenAI API",  "AI chat, store analysis, ROAS opt.",   "Pay-as-you-go",          "~$0.44–$6.50/user/mo",     "Usage-based, scales with users"),
    ("Meta API",    "Campaign management",                  "Free",                   "Free",                     "N/A"),
    ("Google Ads API","Campaign management",                "Free",                   "Free",                     "N/A"),
    ("Shopify API", "Product sync",                         "Free (Partner app)",     "Free",                     "N/A"),
]
svc_colors = [PURPLE, TEAL, GREEN, YELLOW, TEAL, PURPLE, WHITE, WHITE, WHITE]
for i, (svc, role, free, paid, trigger) in enumerate(infra):
    r = 4 + i
    ws4.row_dimensions[r].height = 20
    row_bg = "13162A" if i % 2 == 0 else "0F1120"
    for col, val, col_c in zip(
        ["B","C","D","E","F"],
        [svc, role, free, paid, trigger],
        [svc_colors[i], MUTED, GREEN, YELLOW, WHITE]
    ):
        c = ws4[f"{col}{r}"]
        c.value = val
        c.font = Font(color=col_c, size=10, name="Calibri", bold=(col == "B"))
        c.alignment = left() if col in ["C","D","E","F"] else center()
        c.fill = fill(row_bg)
        c.border = border()

ws4.row_dimensions[13].height = 12

# Monthly infra total table
ws4.merge_cells("B14:F14")
ws4["B14"] = "MONTHLY INFRASTRUCTURE TOTAL (excluding Higgsfield + OpenAI variable costs)"
ws4["B14"].font = Font(bold=True, color=PURPLE, size=11, name="Calibri")
ws4["B14"].alignment = left()
ws4["B14"].fill = fill("100A1A")

hdr(ws4, 15, ["B","C","D","E","F"],
    ["Scale", "Customers", "Vercel", "n8n", "Total Fixed"])

infra_scale = [
    ("Launch",    "0–25",    "$0",  "$7",   "$7"),
    ("Early",     "25–100",  "$20", "$10",  "$30"),
    ("Growth",    "100–500", "$20", "$15",  "$35"),
    ("Scale",     "500+",    "$20", "$40",  "$60"),
]
for i, (stage, cust, vercel, n8n, total) in enumerate(infra_scale):
    r = 16 + i
    ws4.row_dimensions[r].height = 20
    row_bg = "13162A" if i % 2 == 0 else "0F1120"
    for col, val, col_c in zip(
        ["B","C","D","E","F"],
        [stage, cust, vercel, n8n, total],
        [TEAL, WHITE, YELLOW, YELLOW, GREEN]
    ):
        c = ws4[f"{col}{r}"]
        c.value = val
        c.font = Font(color=col_c, size=10, name="Calibri", bold=(col in ["B","F"]))
        c.alignment = center()
        c.fill = fill(row_bg)
        c.border = border()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Switch Analysis
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📈 PAYG vs ULTRA")
ws5.sheet_view.showGridLines = False

col_w5 = [2, 16, 16, 16, 16, 16, 16, 16, 2]
for i, w in enumerate(col_w5):
    ws5.column_dimensions[get_column_letter(i+1)].width = w

for row in ws5.iter_rows(min_row=1, max_row=50, min_col=1, max_col=9):
    for cell in row:
        cell.fill = fill(NAVY)

ws5.merge_cells("B1:H1")
ws5["B1"] = "PAYG vs ULTRA ANNUAL — BREAK-EVEN ANALYSIS"
ws5["B1"].font = Font(bold=True, color=TEAL, size=14, name="Calibri")
ws5["B1"].alignment = center()
ws5.row_dimensions[1].height = 36
ws5.row_dimensions[2].height = 10

ws5.merge_cells("B3:H3")
ws5["B3"] = "At what customer count does ULTRA annual ($99/mo flat) beat PAYG?"
ws5["B3"].font = Font(bold=True, color=YELLOW, size=11, name="Calibri")
ws5["B3"].alignment = left()
ws5["B3"].fill = fill("11142B")

hdr(ws5, 4, ["B","C","D","E","F","G","H"],
    ["Customers", "Mix (S/G/Sc)", "Videos/mo", "PAYG Cost", "ULTRA Cost", "Savings", "Verdict"])

switch_scenarios = [1, 2, 3, 4, 5, 8, 10, 15, 20, 25]

for i, n in enumerate(switch_scenarios):
    r = 5 + i
    ws5.row_dimensions[r].height = 20
    s = round(n*0.60); g = round(n*0.30); sc = n - s - g
    total_vids = s*15 + g*75 + sc*350
    payg_cost = total_vids * 0.28
    ultra_cost = 99.0
    savings = payg_cost - ultra_cost
    verdict = "✅ Switch to ULTRA" if savings > 0 else "💤 Stay PAYG"
    mix = f"{s}/{g}/{sc}"
    row_bg = "001A00" if savings > 0 else ("13162A" if i % 2 == 0 else "0F1120")
    verdict_color = GREEN if savings > 0 else MUTED
    for col, val, col_c in zip(
        ["B","C","D","E","F","G","H"],
        [str(n), mix, str(total_vids), money(payg_cost), "$99", money(max(0,savings)), verdict],
        [WHITE, MUTED, WHITE, YELLOW, GREEN, GREEN if savings > 0 else RED, verdict_color]
    ):
        c = ws5[f"{col}{r}"]
        c.value = val
        c.font = Font(color=col_c, size=10, name="Calibri", bold=(col in ["G","H"]))
        c.alignment = center()
        c.fill = fill(row_bg)
        c.border = border()

ws5.row_dimensions[15].height = 12
ws5.merge_cells("B16:H16")
ws5["B16"] = "💡  Crossover: ~4 customers (60/30/10 mix). At 4 customers PAYG ≈ $95 vs ULTRA $99 — near breakeven. At 5+ customers ULTRA wins."
ws5["B16"].font = Font(color=TEAL, size=10, name="Calibri", bold=True)
ws5["B16"].alignment = left()
ws5["B16"].fill = fill("001215")


# ═══════════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════════
out_path = r"C:\Users\jamal settah\Desktop\Ephermal_Financials.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
